import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# ✅ 네이버 지도 검색 URL
BASE_URL = "https://map.naver.com/p/search/"

# ✅ 크롬 드라이버 옵션 설정
options = webdriver.ChromeOptions()
# options.add_argument("--headless")  # 결과를 직접 확인하고 싶다면 주석 해제하지 말 것
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

# ✅ 크롬 드라이버 실행
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# ✅ 원본 Excel 파일 읽기 (가게.xlsx)
input_excel = "가게2.xlsx"
try:
    df = pd.read_excel(input_excel)
except FileNotFoundError:
    print(f"❌ 파일 '{input_excel}'을 찾을 수 없습니다.")
    driver.quit()
    exit()

# ✅ 필수 컬럼 확인 (가게.xlsx에는 아래 컬럼들이 존재해야 함)
required_columns = ["시설명", "카테고리", "위도", "경도"]
for col in required_columns:
    if col not in df.columns:
        print(f"❌ 필수 컬럼 '{col}' 이(가) 없습니다.")
        driver.quit()
        exit()

# ✅ 출력 파일 경로 및 헤더 (원본 컬럼 + '네이버플레이스ID')
output_excel = "가게2_업데이트.xlsx"
# 출력 파일이 없으면 새로 생성해서 헤더 기록
if not os.path.exists(output_excel):
    wb = Workbook()
    ws = wb.active
    header = required_columns + ["네이버플레이스ID"]
    ws.append(header)
    wb.save(output_excel)
    print(f"✅ 출력 파일 '{output_excel}'을 생성하였습니다.")

# ✅ 각 행에 대해 네이버 지도에서 플레이스 ID 수집 후, 네이버플레이스ID가 존재하는 경우에만 추가
total = len(df)
for idx, row in df.iterrows():
    facility = row["시설명"]  # 시설명으로 검색
    print(f"🔍 '{facility}' 검색 중... ({idx+1}/{total})")
    
    search_url = BASE_URL + facility
    driver.get(search_url)
    time.sleep(5)  # URL 변화 반영을 위한 대기

    naver_place_id = None  # 기본값 None (검색 결과가 없으면 null)
    try:
        # ✅ iframe이 로드될 때까지 대기 후 전환
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "searchIframe"))
        )
        driver.switch_to.frame("searchIframe")
        time.sleep(1)  # 내부 로딩 대기

        # ✅ 첫 번째 검색 결과 요소 찾기 (CSS 선택자)
        place_element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a.ApCpt.k4f_J"))
        )
        if place_element:
            print(f"✅ '{facility}' 검색 결과 발견!")
            # 첫 번째 결과 클릭
            driver.execute_script("arguments[0].click();", place_element)
            time.sleep(1)  # 상세 페이지 로딩 대기

            # ✅ iframe 해제 후 현재 URL에서 플레이스 ID 추출
            driver.switch_to.default_content()
            time.sleep(1)
            current_url = driver.current_url
            if "/place/" in current_url:
                naver_place_id = current_url.split("/place/")[-1].split("?")[0]
            else:
                print(f"❌ '{facility}'의 URL에서 ID 추출 실패.")
        else:
            print(f"❌ '{facility}' 검색 결과 없음.")
    except Exception as e:
        print(f"❌ '{facility}' 검색 중 오류 발생: {e}")
        naver_place_id = None

    # 네이버플레이스ID가 존재할 때만 출력 파일에 추가
    if naver_place_id is not None:
        # ✅ 새로운 행 데이터 생성 (원본 데이터 + 네이버플레이스ID)
        row_data = [row[col] for col in required_columns] + [naver_place_id]
    
        # ✅ 출력 파일을 열어 해당 행 데이터 append & 저장
        wb = load_workbook(output_excel)
        ws = wb.active
        ws.append(row_data)
        wb.save(output_excel)
        print(f"✅ '{facility}'의 네이버플레이스ID: {naver_place_id} 추가됨.")
    else:
        print(f"⚠ '{facility}'의 네이버플레이스ID가 None이므로 추가하지 않습니다.")
    
    # ✅ 다음 검색을 위해 iframe 해제 및 대기
    driver.switch_to.default_content()
    time.sleep(1)

# ✅ 크롬 드라이버 종료
driver.quit()
print(f"✅ 모든 작업 완료! 업데이트 파일: {output_excel}")
